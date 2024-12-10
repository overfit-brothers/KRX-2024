import os
import time
import json
import argparse
from typing import List, Dict
from unsloth import FastLanguageModel
from transformers import TextStreamer, AutoTokenizer, AutoModelForCausalLM
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def load_model_and_tokenizer(model_name: str, mode='normal'):
    if mode == 'normal':
        model, tokenizer = AutoModelForCausalLM.from_pretrained(
            model_name=model_name,
            max_seq_length=1024,
            token=os.getenv("HF_TOKEN")
        )
        return model, tokenizer
    elif mode == 'unsloth':
        model, tokenizer = FastLanguageModel.from_pretrained(
            model_name=model_name,
            max_seq_length=1024,
            dtype=None,
            load_in_4bit=False,
            token=os.getenv("HF_TOKEN")
        )
        return model, tokenizer

def get_test_questions() -> Dict[str, List[str]]:
    
    with open('script/KRX/questions.json', 'r', encoding='utf-8') as file:
        questions = json.load(file)
    return questions

# 모든 질문을 리스트로 얻고 싶다면 다음과 같이 사용할 수 있습니다:
def get_all_questions() -> List[str]:
    questions_dict = get_test_questions()
    return [q for category in questions_dict.values() for q in category]

# 특정 카테고리의 질문만 얻고 싶다면 다음과 같이 사용할 수 있습니다:
def get_questions_by_category(category: str) -> List[str]:
    questions_dict = get_test_questions()
    return questions_dict.get(category, [])

def generate_answer(model, tokenizer, text_streamer, question: str) -> Dict[str, str]:
    messages = [
        {"role": "user", "content": question}
    ]
    input_text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer(input_text, return_tensors="pt").to("cuda")

    start_time = time.time()
    answer = model.generate(**inputs,streamer=text_streamer, max_new_tokens = 1024, use_cache = True)
    end_time = time.time()
    
    return {
        "question": question,
        "answer": tokenizer.decode(answer[0], skip_special_tokens=True).split('\nassistant\n')[-1],
        "output_time": end_time - start_time
    }

def save_to_json(data: List[Dict[str, str]], filename: str):
    with open(filename, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)

def model_run(model_name,mode):
    model, tokenizer = load_model_and_tokenizer(model_name,mode=mode)

    FastLanguageModel.for_inference(model)
    text_streamer = TextStreamer(tokenizer)

    questions = get_all_questions()
    responses = []

    for question in questions:
        response = generate_answer(model, tokenizer, text_streamer, question)
        responses.append(response)
        
    return responses

def save_to_xlsx(tuned_responses, backbone_responses, tuned_model_name, backbone_model_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["질문", f"{tuned_model_name}\n답변", f"{backbone_model_name}\n답변", 
               f"{tuned_model_name}\n처리속도(초)", f"{backbone_model_name}\n처리속도(초)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    # Data
    for row, (tuned, backbone) in enumerate(zip(tuned_responses, backbone_responses), start=2):
        ws.cell(row=row, column=1, value=tuned["question"]).alignment = wrap_alignment
        ws.cell(row=row, column=2, value=tuned["answer"]).alignment = wrap_alignment
        ws.cell(row=row, column=3, value=backbone["answer"]).alignment = wrap_alignment
        ws.cell(row=row, column=4, value=round(tuned["output_time"], 2)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=round(backbone["output_time"], 2)).alignment = Alignment(horizontal="center")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    wb.save(filename)

def main(backbone_model, tuned_model, mode='unsloth'):
    file_name = "Qwen2.5-7B-Instruct"
    backbone_model_name = backbone_model
    tuned_model_name = tuned_model
    
    backbone_responses = model_run(backbone_model_name,mode=mode)
    tuned_responses = model_run(tuned_model_name,mode=mode)
    
    if not os.path.exists("model_results"):
        os.mkdir("model_results")
    # Generate Excel file
    excel_filename = f"model_results/{file_name}.xlsx"
    save_to_xlsx(tuned_responses, backbone_responses, tuned_model_name, backbone_model_name, excel_filename)

    print(f"비교 결과가 {excel_filename}에 Excel 형식으로 저장되었습니다.")

    # JSON 결과저장
    json_filename = f"model_results/{file_name}.json"
    full_results = {
        "tuned_model": {
            "name": tuned_model_name,
            "responses": tuned_responses
        },
        "backbone_model": {
            "name": backbone_model_name,
            "responses": backbone_responses
        }
    }
    save_to_json(full_results, json_filename)
    print(f"전체 결과가 {json_filename}에 JSON 형식으로 저장되었습니다.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run model comparison")
    parser.add_argument("--backbone_model","-b", type=str, help="Name of the backbone model")
    parser.add_argument("--tuned_model","-t", type=str, help="Name of the tuned model")
    parser.add_argument("--mode", type=str, default="unsloth", help="Mode for model loading (default: unsloth)")
    
    args = parser.parse_args()
    
    main(args.backbone_model, args.tuned_model, args.mode)